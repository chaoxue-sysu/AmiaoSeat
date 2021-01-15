# @Date: 2021/01/12
# @Author: Xue Chao
# @Description:
import tkinter.filedialog
import tkinter.messagebox
from tkinter import *
from tkinter.font import Font
from PIL import Image,ImageTk
import os,threading
import openpyxl
import random
import time



curWidth,curHight=640,280

def new_win():
    tkinter.messagebox.showinfo(title='提示', message='运行成功！点击 “确定” 查看结果！')
    check()

def set_priority(seats,students_num):
    col_row=seats
    cols=sorted(col_row.keys())
    print(f'{len(cols)} col seats')
    print(cols)
    all_seats=0
    for col in cols:
        all_seats+=len(col_row[col])
    priors=[]
    skip=False
    for i in range(min(cols),max(cols)+1):
        if skip or not cols.__contains__(i):
            i+=1
            skip=False
            continue
        priors.append(i)
        skip=True
    if all_seats/2<students_num:
        col_pro={}
        for i in cols:
            if not priors.__contains__(i):
                idx=cols.index(i)
                try:
                    f=cols[idx]-cols[idx-1]
                except:
                    f=2
                try:
                    n=cols[idx+1]-cols[idx]
                except:
                    n=2
                col_pro[i]=n+f
        for x in sorted(col_pro.keys(),key=lambda x:(-col_pro[x],x)):
            priors.append(x)
    return priors

def seat(path,out_path):
    workbook=openpyxl.load_workbook(path)
    sh1=workbook.worksheets[1]
    seat_sh=workbook.worksheets[0]
    names={}
    name_idx=[]
    for i in range(sh1.max_row):
        for j in range(sh1.max_column):
            na=str(sh1.cell(i+1,j+1).value).strip()
            if na.__contains__('姓名'):
                name_idx.append([i,j])
    for r,j in name_idx:
        for i in range(r+1,sh1.max_row):
            try:
                name=sh1.cell(i+1,j+1).value.strip()
            except:
                name=''
            if name=='':
                continue
            names[name]=[i+1,j-2+1]
    nums={}
    for i in range(3,seat_sh.max_row):
        for j in range(seat_sh.max_column):
            try:
                snum=int(seat_sh.cell(i+1,j+1).value)
                nums[snum]=[i+1,j+1]
            except:
                continue
    if len(nums)/2<len(names):
        print(f'warning: {path} not ok!: info: {len(nums)} seats vs {len(names)} students')
    else:
        print(f'info:{path}  {len(nums)} seats vs {len(names)} students')

    col_row={}
    for seat in nums.keys():
        r,c=nums[seat]
        if not col_row.__contains__(c):
            col_row[c]=[]
        col_row[c].append([r,seat])
    priors=set_priority(col_row,len(names))
    print(priors)
    i=-1
    finish=False
    stu_names=sorted(names.keys(),key=lambda x:random.random())
    for col in priors:
        if finish:
            break
        for row,sid in col_row[col]:
            if finish:
                break
            i+=1
            if i>=len(stu_names):
                finish=True
                continue
            seat_sh.cell(row+1,col,stu_names[i])
            sh1.cell(*names[stu_names[i]],sid)
    workbook.save(out_path)


def do():
    out_dir=str(entry2.get()).strip()
    try:
        files=in_files
    except:
        files=[]
    if len(files)==0 or out_dir=='':
        tkinter.messagebox.showerror(title='错误', message='您还没有选择输入文件或者输出路径哦！')
        return
    if not os.path.isdir(out_dir):
        os.makedirs(out_dir)
    for p in files:
        path=p
        print(path)
        if not os.path.exists(path):
            continue
        np='.'.join(os.path.basename(p).split('.')[:-1]+['amiu','xlsx'])
        out_path=f'{out_dir}/{np}'
        seat(path,out_path)
    new_win()

def select_multi():
    files=tkinter.filedialog.askopenfilenames(title='选择文件（可多选）', filetypes=[
        ("Excel format", ".xlsx"),
    ])
    x=';'.join(list(files))
    global in_files
    in_files=list(files)
    entry1.delete(0,END)
    entry1.insert(0,x)

def select_dir():
    dir=tkinter.filedialog.askdirectory()
    entry2.delete(0,END)
    entry2.insert(0,dir)

def check():
    os.startfile(entry2.get())
    # tkinter.filedialog.askopenfilename()

def main():
    # return
    global entry1,entry2
    global rootMSCT,root,finish_load
    finish_load=False
    rootMSCT=Tk()
    rootMSCT.attributes('-alpha',0)
    # root.overrideredirect(True)
    # global curWidth,curHeight
    # curWidth,curHight=640,280
    size_xy = get_pos(curWidth,curHight,rootMSCT)
    rootMSCT.geometry(size_xy)

    root=Toplevel()

    tMain=threading.Thread(target=showWelcome)
    tMain.start()
    t1=threading.Thread(target=closeWelcome)
    t1.start()

    rootMSCT.title('阿喵排排座 V1.0')
    entry1=Entry(rootMSCT,width=50,font=Font(family='Times', size=10))
    entry2=Entry(rootMSCT,width=10,font=Font(family='Times', size=10))
    rootMSCT.iconbitmap('logo.ico')
    lablet=Label(rootMSCT,text='阿喵排排座 V1.0',font=Font(family='Times', size=24, weight='bold'),fg='black')
    lable1=Label(rootMSCT,text='输入.xlsx文件（可多选）',font=Font(family='Times', size=12, weight='bold'),width=25)
    sele_btn1=Button(rootMSCT,command=select_multi,text='选择文件',font=Font(family='Times', size=10, weight='bold'),bg='pink',fg='black',width=8)

    lable2=Label(rootMSCT,text='输出目录',font=Font(family='Times', size=12, weight='bold'),width=10)
    sele_btn2=Button(rootMSCT,command=select_dir,text='选择目录',font=Font(family='Times', size=10, weight='bold'),bg='pink',fg='black',width=8)

    btn=Button(rootMSCT,command=do,text='点击开始排座',font=Font(family='Times', size=15, weight='bold'),bg='pink',fg='black',width=20)
    lablet.grid(row=0,column=0,pady=35,sticky=N+S+E+W,columnspan=3)
    lable1.grid(row=1,column=0,sticky=N+S+E+W)
    entry1.grid(row=1,column=1,sticky=N+S+E+W)
    sele_btn1.grid(row=1,column=2,sticky=N+S+E+W)
    lable2.grid(row=2,column=0,sticky=N+S+E+W, pady=5)
    entry2.grid(row=2,column=1,sticky=N+S+E+W, pady=5)
    sele_btn2.grid(row=2,column=2,sticky=N+S+E+W, pady=5)
    btn.grid(row=3,column=0,columnspan=3,sticky=N+S,pady=30)

    print('finish load main')
    finish_load=True
    rootMSCT.mainloop()

def get_pos(w,h,wind):
    scn_w, scn_h = wind.maxsize()
    cen_x = (scn_w - w) / 2
    cen_y = 0.8*(scn_h - h) / 2
    size_xy = '%dx%d+%d+%d' % (w, h, cen_x, cen_y)
    return size_xy

def showWelcome():
    #得到屏幕高度
    root.overrideredirect(True)
    root.attributes("-alpha", 1)#窗口透明度（1为不透明，0为全透明）
    #设置窗口位于屏幕中部
    xw,xh=600,240
    root.geometry(get_pos(xw,xh,root))
    root['bg']='white'
    #插入欢迎图片，可以是logo
    logo='logo.png'
    if os.path.exists(logo):
        print("Lib/img exist")
        photo = Image.open(logo)
        photo=photo.resize((160,160))
        bm = ImageTk.PhotoImage(photo)
        lb_welcomelogo = Label(root, image = bm,bg='white')
        lb_welcomelogo.bm = bm
        lb_welcomelogo.place(x=40, y=40)
    #插入文字，可以显示开发者或出处
    lb_welcometext = Label(root, text = 'AmiaoSeat V1.0',
                           fg='lightgray',font=Font(family='Times', size=28, weight='bold'),bg='white')
    lb_welcometext.place(x=220, y=35,width=300,height=100)
    canvas = Canvas(root, width=300, height=18, bg="white")
    canvas.place(x=220, y=130)
    cpr = Label(root, text = 'Copyright © 2021 MiuMiu',
                           fg='lightgray',font=Font(family='Times', size=10),bg='white')
    cpr.place(x=208, y=160,width=300,height=50)
    fill_line = canvas.create_rectangle(1.5, 1.5, 0, 23, width=0, fill="pink")
    x = 150  # 未知变量，可更改
    n = 465 / x  # 465是矩形填充满的次数
    for i in range(x):
        n = n + 465 / x
        canvas.coords(fill_line, (0, 0, n, 60))
        root.update()
        time.sleep(0.01)

def closeWelcome():
    # # #设置欢迎页停留时间
    # for i in range(5):
    #     # rootMSCT.attributes("-alpha", 0)#窗口透明度
    #     time.sleep(1)
    time.sleep(2)
    while not finish_load:
        continue
    rootMSCT.attributes("-alpha", 1)#窗口透明度
    rootMSCT.attributes('-topmost',True)
    root.destroy()

def app():
    main()


if __name__=='__main__':
    app()

